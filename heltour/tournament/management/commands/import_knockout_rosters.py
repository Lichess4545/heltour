"""Import team rosters and the first-round bracket for a knockout tournament.

Two roster input shapes are supported:

- ``--lineups <file>`` - one consolidated workbook with a header row
  (``team_name, board_number, lichess_username, fide_name, fide_rating,
  fide_id`` and optionally ``role``) and one player per row across all teams.
- ``--rosters-dir <dir>`` - one workbook per team (the per-team files produced
  by ``scripts/export_bracket_rosters.py``).

And two pairings input shapes (auto-detected):

- a clean ``Pairings`` sheet with ``match_number, team_a, team_b``; or
- the human bracket layout (``Duel N`` blocks with team names that carry a
  trailing kick-off time, e.g. ``Saint Louis University 10:00``).

The full team list comes from the pairings file. A team that appears in the
bracket but has no lineup is created **empty** (no players) - that is expected
while rosters are still trickling in. Round-1 ``TeamPairing`` rows are only
created when ``--create-pairings`` is passed (and then every team must have a
roster).

Because captains hand-edit these files, the importer is defensive: it parses
and validates everything, normalises messy values (usernames pasted as
``@name`` or profile URLs, FIDE ids with stray spaces, free-text roles, team
names with trailing times), and reports *all* problems at once. Nothing is
written unless validation is clean.

Player reuse: players are matched to existing ``Player`` rows by
``lichess_username`` (case-insensitive) first, then by ``fide_id`` as a
fallback. Only when neither matches is a new ``Player`` created.

Examples:

    # Validate the captain-submitted spreadsheets - no DB writes.
    python manage.py import_knockout_rosters --validate-only \\
        --lineups knockout-r1-lineups.xlsx --pairings ko-pairing.xlsx

    # Import into the existing WUCCKNOCKOUT / 2026 season. Teams with no
    # lineup are created empty; round-1 pairings are NOT created yet.
    python manage.py import_knockout_rosters \\
        --league-tag WUCCKNOCKOUT --season-tag 2026 \\
        --lineups knockout-r1-lineups.xlsx --pairings ko-pairing.xlsx
"""

import os
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

ROSTER_HEADER_KEY = "lichess_username"
MAX_BOARD_NUMBER = 12  # heltour BOARD_NUMBER_OPTIONS tops out at 12
USERNAME_RE = re.compile(r"^[\w-]{2,30}$")
DUEL_RE = re.compile(r"duel\s*(\d+)", re.IGNORECASE)
# Team names in the bracket layout carry a trailing kick-off time, e.g.
# "Saint Louis University 10:00" or "Monash Univercity 10:00 +1".
TIME_TAIL_RE = re.compile(r"\s+\d{1,2}:\d{2}(?:\s*\+\d+)?\s*$")

CAPTAIN_ROLES = {"captain", "c", "(c)", "capt", "cap"}
VICE_CAPTAIN_ROLES = {
    "vice_captain",
    "vice-captain",
    "vice captain",
    "vice",
    "vc",
    "v",
    "(v)",
}


class Command(BaseCommand):
    help = "Import knockout team rosters + first-round bracket from xlsx files"

    def add_arguments(self, parser):
        parser.add_argument("--league-tag", help="League tag to import into")
        parser.add_argument(
            "--create-league",
            action="store_true",
            help="Create the league (knockout-multi, team) if it does not exist",
        )
        parser.add_argument("--league-name", help="League name (used with --create-league)")
        parser.add_argument("--season-tag", help="Season tag to import into")
        parser.add_argument(
            "--season-name", help="Season name (defaults to the season tag)"
        )
        parser.add_argument(
            "--lineups",
            help="Consolidated lineups .xlsx (all teams in one sheet)",
        )
        parser.add_argument(
            "--rosters-dir",
            help="Directory of per-team roster .xlsx files",
        )
        parser.add_argument(
            "--pairings",
            required=True,
            help="Pairings .xlsx (clean table or the 'Duel N' bracket layout)",
        )
        parser.add_argument(
            "--rounds",
            type=int,
            default=2,
            help="Number of rounds, used only when creating a new season (default: 2)",
        )
        parser.add_argument(
            "--boards",
            type=int,
            help="Boards per team (default: an existing season's value, else the "
            "largest board number found in the lineups)",
        )
        parser.add_argument(
            "--match-generation",
            choices=["lockstep", "upfront"],
            default="upfront",
            help=(
                "Multi-match generation mode for the bracket. 'upfront' (default) "
                "lets each bracket play its matches back-to-back; 'lockstep' is "
                "for a fully Litour-run event."
            ),
        )
        parser.add_argument(
            "--create-pairings",
            action="store_true",
            help="Also create the round-1 TeamPairing rows now (requires every "
            "team to have a roster). Off by default.",
        )
        parser.add_argument(
            "--clear-existing",
            action="store_true",
            help="Delete the season's existing teams, season players and bracket first",
        )
        parser.add_argument(
            "--validate-only",
            action="store_true",
            help="Parse and validate the spreadsheets only; make no database changes",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Do all the work in a transaction, then roll it back",
        )

    def handle(self, *args, **options):
        lineups_path = options["lineups"]
        rosters_dir = options["rosters_dir"]
        pairings_path = options["pairings"]

        if bool(lineups_path) == bool(rosters_dir):
            raise CommandError("Pass exactly one of --lineups or --rosters-dir")
        if lineups_path and not os.path.isfile(lineups_path):
            raise CommandError(f"Lineups file not found: {lineups_path}")
        if rosters_dir and not os.path.isdir(rosters_dir):
            raise CommandError(f"Rosters directory not found: {rosters_dir}")
        if not os.path.isfile(pairings_path):
            raise CommandError(f"Pairings file not found: {pairings_path}")

        validate_only = options["validate_only"]
        if not validate_only and (not options["league_tag"] or not options["season_tag"]):
            raise CommandError(
                "--league-tag and --season-tag are required unless --validate-only"
            )

        errors = []
        warnings = []

        # --- Phase 1: parse + format-validate the spreadsheets (no DB) --------
        if lineups_path:
            rosters, perr, pwarn = self._parse_lineups_file(lineups_path)
        else:
            rosters, perr, pwarn = self._parse_rosters(rosters_dir)
        errors += perr
        warnings += pwarn

        ordered_pairs, pair_err = self._parse_pairings(pairings_path)
        errors += pair_err

        cross_err, cross_warn, ordered_pairs, rosters, all_teams = self._cross_validate(
            rosters, ordered_pairs
        )
        errors += cross_err
        warnings += cross_warn

        boards = options["boards"] or self._infer_boards(rosters)
        if boards:
            warnings += self._check_board_counts(rosters, boards)

        # --- Phase 2: resolve players against the DB (reads only) ------------
        plan, rerr, rwarn = self._resolve_players(rosters)
        errors += rerr
        warnings += rwarn

        # --- Report ----------------------------------------------------------
        for warning in warnings:
            self.stdout.write(self.style.WARNING(f"WARNING: {warning}"))
        if errors:
            for error in errors:
                self.stdout.write(self.style.ERROR(f"ERROR: {error}"))
            raise CommandError(
                f"{len(errors)} validation error(s); nothing was imported"
            )

        teams_with_rosters = len(rosters)
        teams_empty = len(all_teams) - teams_with_rosters
        reused = sum(1 for p in plan.values() if p["action"] == "reuse")
        new = sum(1 for p in plan.values() if p["action"] == "create")
        self.stdout.write(
            self.style.SUCCESS(
                f"Validation passed: {len(all_teams)} teams "
                f"({teams_with_rosters} with rosters, {teams_empty} empty), "
                f"{len(plan)} players ({reused} reused, {new} new), "
                f"{len(ordered_pairs)} bracket matches, {len(warnings)} warning(s)"
            )
        )
        if validate_only:
            return

        # --- Phase 3: write ---------------------------------------------------
        with transaction.atomic():
            league = self._resolve_league(options)
            season = self._resolve_season(league, options, boards)

            if options["clear_existing"]:
                self._clear_season(season)
            self._guard_season_empty(season)

            rounds = self._ensure_rounds(season)
            round1 = rounds[0]

            teams_by_name = self._import_teams(season, all_teams, rosters, plan)

            bracket = self._create_bracket(
                season, len(all_teams), options["match_generation"]
            )
            self._create_seedings(bracket, all_teams, teams_by_name)

            from heltour.tournament_core.knockout import get_knockout_stage_name

            round1.knockout_stage = get_knockout_stage_name(len(all_teams))
            round1.save()

            created_pairings = 0
            if options["create_pairings"]:
                empty = [n for n in all_teams if not rosters.get(n)]
                if empty:
                    raise CommandError(
                        "--create-pairings needs every team to have a roster; "
                        "still empty: " + ", ".join(empty)
                    )
                created_pairings = self._create_round1_pairings(
                    round1, ordered_pairs, teams_by_name, bracket
                )

            season.calculate_scores()

            self.stdout.write(
                self.style.SUCCESS(
                    f"Imported {len(teams_by_name)} teams "
                    f"({teams_with_rosters} with rosters, {teams_empty} empty) "
                    f"into {league.tag}/{season.tag}"
                )
            )
            self.stdout.write(f"  Season ID: {season.pk}  (use with generate_random_results)")
            self.stdout.write(
                f"  Bracket: {bracket.bracket_size} teams, "
                f"matches_per_stage={bracket.matches_per_stage}, "
                f"match_generation={bracket.match_generation}"
            )
            if created_pairings:
                self.stdout.write(
                    f"  Round 1 ({round1.knockout_stage}): {created_pairings} pairings created"
                )
            else:
                self.stdout.write(
                    f"  Round 1 ({round1.knockout_stage}): pairings NOT created "
                    "(pass --create-pairings once every roster is in)"
                )
            self.stdout.write(
                f"  Teams page: /{league.tag}/season/{season.tag}/teams/"
            )

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("DRY RUN: rolled back, nothing saved"))

    # ----- phase 1: parsing + format validation -------------------------------

    def _parse_lineups_file(self, path):
        """Parse a consolidated lineups workbook. Returns (rosters, errors, warnings).

        ``rosters`` maps team_name -> list of normalised player-row dicts.
        """
        from openpyxl import load_workbook

        rosters = {}
        errors = []
        warnings = []

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return rosters, [f"{path}: could not open ({exc})"], warnings
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True)) if ws is not None else None
        finally:
            wb.close()
        if rows is None:
            return rosters, [f"{path}: workbook has no active sheet"], warnings

        header_idx, headers = self._find_header(rows)
        if header_idx is None:
            return rosters, [
                f"{path}: no header row found (expected a '{ROSTER_HEADER_KEY}' column)"
            ], warnings
        col = {name: i for i, name in enumerate(headers) if name}
        for required in ("team_name", "board_number"):
            if required not in col:
                errors.append(f"{path}: header is missing the '{required}' column")
        if errors:
            return rosters, errors, warnings

        grouped = {}
        for sheet_row, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            get = self._cell_getter(col, row)
            team_raw = get("team_name")
            if self._row_is_blank(get):
                continue
            where = f"{path} row {sheet_row}"
            team_name = str(team_raw).strip() if team_raw not in (None, "") else ""
            if not team_name:
                errors.append(f"{where}: row has no team_name")
                continue
            parsed = self._parse_player_row(get, where, errors, warnings)
            if parsed is None:
                continue
            grouped.setdefault(team_name, []).append(parsed)

        for team_name, team_rows in grouped.items():
            self._validate_team_boards(team_name, team_rows, errors, warnings)
            team_rows.sort(key=lambda r: r["board_number"])
            rosters[team_name] = team_rows

        if not rosters and not errors:
            errors.append(f"{path}: no player rows found")
        return rosters, errors, warnings

    def _parse_rosters(self, rosters_dir):
        """Parse a directory of per-team roster .xlsx files."""
        from openpyxl import load_workbook

        rosters = {}
        errors = []
        warnings = []

        files = sorted(
            f for f in os.listdir(rosters_dir)
            if f.lower().endswith(".xlsx") and not f.startswith("~$")
        )
        if not files:
            errors.append(f"No .xlsx roster files found in {rosters_dir}")
            return rosters, errors, warnings

        for filename in files:
            path = os.path.join(rosters_dir, filename)
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
            except Exception as exc:
                errors.append(f"{filename}: could not open ({exc})")
                continue
            try:
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True)) if ws is not None else None
                sheet_title = ws.title if ws is not None else None
            finally:
                wb.close()
            if rows is None:
                errors.append(f"{filename}: workbook has no active sheet")
                continue

            team_rows, team_name, file_errors, file_warnings = self._parse_one_roster(
                filename, sheet_title, rows
            )
            errors += file_errors
            warnings += file_warnings
            if file_errors or team_name is None:
                continue
            if team_name in rosters:
                errors.append(
                    f"{filename}: duplicate team name {team_name!r} "
                    "(already seen in another roster file)"
                )
                continue
            rosters[team_name] = team_rows

        return rosters, errors, warnings

    def _parse_one_roster(self, filename, sheet_title, rows):
        """Parse + validate a single per-team roster sheet.

        Returns (team_rows, team_name, errors, warnings).
        """
        errors = []
        warnings = []

        header_idx, headers = self._find_header(rows)
        if header_idx is None:
            errors.append(
                f"{filename}: no header row found (expected a '{ROSTER_HEADER_KEY}' column)"
            )
            return [], None, errors, warnings
        col = {name: i for i, name in enumerate(headers) if name}
        if "board_number" not in col:
            errors.append(f"{filename}: header is missing the 'board_number' column")
            return [], None, errors, warnings

        team_rows = []
        team_names_seen = set()
        for sheet_row, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            get = self._cell_getter(col, row)
            if self._row_is_blank(get):
                continue
            where = f"{filename} row {sheet_row}"
            parsed = self._parse_player_row(get, where, errors, warnings)
            if parsed is None:
                continue
            row_team = get("team_name")
            if row_team and str(row_team).strip():
                team_names_seen.add(str(row_team).strip())
            team_rows.append(parsed)

        if not team_rows:
            errors.append(f"{filename}: no player rows found below the header")
            return [], None, errors, warnings

        if len(team_names_seen) > 1:
            warnings.append(
                f"{filename}: rows disagree on team_name ({sorted(team_names_seen)}); "
                f"using {sorted(team_names_seen)[0]!r}"
            )
        if team_names_seen:
            team_name = sorted(team_names_seen)[0]
        else:
            team_name = (sheet_title or os.path.splitext(filename)[0]).strip()
            warnings.append(
                f"{filename}: no team_name column value; using {team_name!r}"
            )

        self._validate_team_boards(team_name, team_rows, errors, warnings)
        team_rows.sort(key=lambda r: r["board_number"])
        return team_rows, team_name, errors, warnings

    def _parse_player_row(self, get, where, errors, warnings):
        """Normalise + validate one player row. Returns a row dict or None."""
        username, uwarn = self._normalize_username(get("lichess_username"))
        if uwarn:
            warnings.append(f"{where}: {uwarn}")
        fide_id, fwarn = self._normalize_fide_id(get("fide_id"))
        if fwarn:
            warnings.append(f"{where}: {fwarn}")
        if not username and not fide_id:
            errors.append(f"{where}: row has neither a lichess_username nor a fide_id")
            return None

        board_number = self._parse_int(get("board_number"))
        if board_number is None:
            errors.append(
                f"{where}: board_number {get('board_number')!r} is not a number"
            )
            return None
        if board_number < 1 or board_number > MAX_BOARD_NUMBER:
            errors.append(
                f"{where}: board_number {board_number} is out of range "
                f"(1-{MAX_BOARD_NUMBER})"
            )
            return None

        rating_raw = get("fide_rating")
        fide_rating = self._parse_int(rating_raw)
        if rating_raw not in (None, "") and fide_rating is None:
            warnings.append(f"{where}: ignoring unparseable fide_rating {rating_raw!r}")

        role, role_warn = self._normalize_role(get("role"))
        if role_warn:
            warnings.append(f"{where}: {role_warn}")

        return {
            "where": where,
            "board_number": board_number,
            "lichess_username": username,
            "fide_id": fide_id,
            "fide_rating": fide_rating,
            "role": role,
        }

    def _validate_team_boards(self, team_name, team_rows, errors, warnings):
        boards_seen = [r["board_number"] for r in team_rows]
        duplicate = sorted({b for b in boards_seen if boards_seen.count(b) > 1})
        if duplicate:
            errors.append(
                f"{team_name}: duplicate board number(s): "
                + ", ".join(str(b) for b in duplicate)
            )
        elif sorted(boards_seen) != list(range(1, len(boards_seen) + 1)):
            warnings.append(
                f"{team_name}: board numbers are not 1..{len(boards_seen)} "
                f"(got {sorted(boards_seen)})"
            )
        captains = [r for r in team_rows if r["role"] == "captain"]
        if len(captains) > 1:
            warnings.append(f"{team_name}: {len(captains)} captains marked (expected 1)")

    def _parse_pairings(self, path):
        """Parse the pairings workbook, auto-detecting the layout.

        Returns (ordered_pairs, errors) with ordered_pairs sorted by match
        number: [(match_number, team_a, team_b), ...].
        """
        from openpyxl import load_workbook

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return [], [f"{path}: could not open ({exc})"]
        try:
            ws = wb["Pairings"] if "Pairings" in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True)) if ws is not None else None
        finally:
            wb.close()

        if rows is None:
            return [], [f"{path}: workbook has no active sheet"]
        if not rows:
            return [], [f"{path}: the sheet is empty"]

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        if {"match_number", "team_a", "team_b"} <= set(header):
            return self._parse_pairings_table(path, rows, header)
        return self._parse_pairings_bracket(path, rows)

    def _parse_pairings_table(self, path, rows, header):
        """Parse the clean ``match_number, team_a, team_b`` layout."""
        errors = []
        col = {name: i for i, name in enumerate(header) if name}
        pairs = []
        seen = set()
        for sheet_row, row in enumerate(rows[1:], start=2):
            get = self._cell_getter(col, row)
            team_a = get("team_a")
            team_b = get("team_b")
            match_raw = get("match_number")
            if all(v in (None, "") for v in (team_a, team_b, match_raw)):
                continue
            where = f"{path} row {sheet_row}"
            match_number = self._parse_int(match_raw)
            if match_number is None:
                errors.append(f"{where}: match_number {match_raw!r} is not a number")
                continue
            if match_number in seen:
                errors.append(f"{where}: duplicate match_number {match_number}")
                continue
            seen.add(match_number)
            if not team_a or not str(team_a).strip():
                errors.append(f"{where}: team_a is blank")
                continue
            if not team_b or not str(team_b).strip():
                errors.append(f"{where}: team_b is blank")
                continue
            pairs.append((match_number, str(team_a).strip(), str(team_b).strip()))

        if not pairs and not errors:
            errors.append(f"{path}: no pairing rows found")
        pairs.sort(key=lambda p: p[0])
        return pairs, errors

    def _parse_pairings_bracket(self, path, rows):
        """Parse the human bracket layout: ``Duel N`` label rows, with the two
        team names (carrying trailing kick-off times) on the row below."""
        errors = []

        # The two team columns are marked by a 'TEAM' header cell.
        team_cols = []
        for row in rows:
            cols = [
                i
                for i, c in enumerate(row)
                if c is not None and str(c).strip().lower() == "team"
            ]
            if len(cols) == 2:
                team_cols = cols
                break
        if len(team_cols) != 2:
            return [], [
                f"{path}: could not find the two 'TEAM' columns in the bracket layout"
            ]

        pairs = []
        seen = set()
        for idx, row in enumerate(rows):
            label = None
            for c in row:
                if c is None:
                    continue
                m = DUEL_RE.search(str(c))
                if m:
                    label = int(m.group(1))
                    break
            if label is None:
                continue
            if idx + 1 >= len(rows):
                errors.append(f"{path}: Duel {label} has no team row after it")
                continue
            team_row = rows[idx + 1]

            def at(i):
                return team_row[i] if i < len(team_row) else None

            a_raw, b_raw = at(team_cols[0]), at(team_cols[1])
            if not a_raw or not str(a_raw).strip():
                errors.append(f"{path}: Duel {label} has a blank team A")
                continue
            if not b_raw or not str(b_raw).strip():
                errors.append(f"{path}: Duel {label} has a blank team B")
                continue
            if label in seen:
                errors.append(f"{path}: duplicate Duel {label}")
                continue
            seen.add(label)
            pairs.append(
                (label, self._strip_time(str(a_raw)), self._strip_time(str(b_raw)))
            )

        if not pairs and not errors:
            errors.append(f"{path}: no 'Duel N' rows found")
        pairs.sort(key=lambda p: p[0])
        return pairs, errors

    def _cross_validate(self, rosters, ordered_pairs):
        """Reconcile lineups against the bracket.

        Returns (errors, warnings, ordered_pairs, rosters_canon, all_teams).
        The full team list comes from the pairings file; ``rosters_canon`` is
        the lineups re-keyed onto those canonical team names.
        """
        errors = []
        warnings = []

        all_teams = []
        for _mn, a, b in ordered_pairs:
            for name in (a, b):
                if name not in all_teams:
                    all_teams.append(name)

        fold_to_canon = {self._fold(n): n for n in all_teams}
        if len(fold_to_canon) != len(all_teams):
            errors.append(
                "pairings file has team names that differ only by case/spacing"
            )
        if ordered_pairs and len(all_teams) != len(ordered_pairs) * 2:
            errors.append(
                f"pairings file has {len(ordered_pairs)} matches but "
                f"{len(all_teams)} distinct teams (is a team listed twice?)"
            )
        if all_teams and not self._is_power_of_2(len(all_teams)):
            errors.append(
                f"{len(all_teams)} teams in the pairings file is not a power "
                "of 2; knockout brackets require a power of 2"
            )

        rosters_canon = {}
        for roster_name, roster_rows in rosters.items():
            canon = fold_to_canon.get(self._fold(roster_name))
            if canon is None:
                errors.append(
                    f"lineup team {roster_name!r} does not match any team in "
                    "the pairings file"
                )
                continue
            if canon in rosters_canon:
                errors.append(f"two lineup entries map to the same team {canon!r}")
                continue
            rosters_canon[canon] = roster_rows

        for name in all_teams:
            if name not in rosters_canon:
                warnings.append(
                    f"{name}: no roster submitted - the team will be created "
                    "with no players"
                )

        return errors, warnings, ordered_pairs, rosters_canon, all_teams

    def _infer_boards(self, rosters):
        counts = [len(rows) for rows in rosters.values() if rows]
        return max(counts) if counts else 0

    def _check_board_counts(self, rosters, boards):
        warnings = []
        for team_name, rows in rosters.items():
            if rows and len(rows) != boards:
                warnings.append(
                    f"{team_name}: has {len(rows)} players but the season has "
                    f"{boards} boards"
                )
        return warnings

    # ----- phase 2: resolve players against the DB (reads only) ---------------

    def _resolve_players(self, rosters):
        """Decide, per roster row, whether to reuse or create a Player.

        Returns (plan, errors, warnings) where plan maps
        (team_name, board_number) -> {action, player, username, fide_id, rating}.
        """
        from heltour.tournament.models import Player

        plan = {}
        errors = []
        warnings = []
        used = {}  # dedup key -> team_name (catches a player on two teams)

        for team_name, rows in rosters.items():
            for row in rows:
                where = row["where"]
                username = row["lichess_username"]
                fide_id = row["fide_id"]

                player = None
                if username:
                    player = Player.objects.filter(
                        lichess_username__iexact=username
                    ).first()
                if player is None and fide_id:
                    matches = list(Player.objects.filter(fide_id=fide_id))
                    if len(matches) == 1:
                        player = matches[0]
                        if username and not self._same_username(
                            player.lichess_username, username
                        ):
                            warnings.append(
                                f"{where}: lichess_username {username!r} not found; "
                                f"matched existing player {player.lichess_username!r} "
                                f"by FIDE id {fide_id}"
                            )
                    elif len(matches) > 1:
                        errors.append(
                            f"{where}: FIDE id {fide_id} matches {len(matches)} "
                            "existing players ("
                            + ", ".join(m.lichess_username for m in matches)
                            + "); resolve manually"
                        )
                        continue

                if player is not None:
                    action = "reuse"
                    dedup_key = player.pk
                else:
                    if not username:
                        errors.append(
                            f"{where}: no lichess_username and FIDE id {fide_id} "
                            "did not match an existing player - cannot create one"
                        )
                        continue
                    if not USERNAME_RE.match(username):
                        warnings.append(
                            f"{where}: {username!r} does not look like a lichess "
                            "username; creating a new player anyway"
                        )
                    action = "create"
                    dedup_key = ("new", username.lower())

                if dedup_key in used:
                    errors.append(
                        f"{where}: player {username or fide_id} also appears on "
                        f"team {used[dedup_key]!r}"
                    )
                    continue
                used[dedup_key] = team_name

                plan[(team_name, row["board_number"])] = {
                    "action": action,
                    "player": player,
                    "username": username,
                    "fide_id": fide_id,
                    "rating": row["fide_rating"],
                }

        return plan, errors, warnings

    # ----- phase 3: database writes -------------------------------------------

    def _resolve_league(self, options):
        from heltour.tournament.models import League

        tag = options["league_tag"]
        league = League.objects.filter(tag=tag).first()
        if league is not None:
            if league.competitor_type != "team":
                raise CommandError(f"League {tag!r} is not a team league")
            if league.pairing_type != "knockout-multi":
                raise CommandError(
                    f"League {tag!r} has pairing_type {league.pairing_type!r}; "
                    "this importer needs 'knockout-multi'"
                )
            return league

        if not options["create_league"]:
            raise CommandError(
                f"League {tag!r} does not exist (pass --create-league to create it)"
            )
        league = League.objects.create(
            name=options["league_name"] or tag,
            tag=tag,
            theme="blue",
            rating_type="classical",
            competitor_type="team",
            pairing_type="knockout-multi",
            knockout_seeding_style="traditional",
            knockout_games_per_match=1,
            is_active=True,
        )
        self.stdout.write(f"Created league {league.tag!r}")
        return league

    def _resolve_season(self, league, options, boards):
        from heltour.tournament.models import Season

        tag = options["season_tag"]
        season = Season.objects.filter(league=league, tag=tag).first()
        if season is not None:
            if season.boards is None:
                season.boards = boards
                season.save()
            elif boards and season.boards != boards:
                self.stdout.write(
                    self.style.WARNING(
                        f"Season boards={season.boards} but lineups imply {boards}; "
                        "using the season's value"
                    )
                )
            return season

        season = Season.objects.create(
            league=league,
            name=options["season_name"] or tag,
            tag=tag,
            rounds=options["rounds"],
            boards=boards,
            start_date=timezone.now(),
            is_active=True,
        )
        self.stdout.write(f"Created season {season.tag!r} (boards={boards})")
        return season

    def _clear_season(self, season):
        from heltour.tournament.models import KnockoutBracket, SeasonPlayer

        team_count = season.team_set.count()
        season.team_set.all().delete()  # cascades to members + pairings
        KnockoutBracket.objects.filter(season=season).delete()
        SeasonPlayer.objects.filter(season=season).delete()
        self.stdout.write(f"Cleared {team_count} existing teams from {season.tag}")

    def _guard_season_empty(self, season):
        if season.team_set.exists():
            raise CommandError(
                f"Season {season.tag!r} already has teams; rerun with "
                "--clear-existing or use a fresh season tag"
            )

    def _ensure_rounds(self, season):
        from heltour.tournament.models import Round

        rounds = []
        for number in range(1, (season.rounds or 1) + 1):
            start = season.start_date + season.round_duration * (number - 1)
            round_obj, _created = Round.objects.get_or_create(
                season=season,
                number=number,
                defaults={
                    "start_date": start,
                    "end_date": start + season.round_duration,
                },
            )
            rounds.append(round_obj)
        return rounds

    def _import_teams(self, season, all_teams, rosters, plan):
        import reversion

        from heltour.tournament.models import (
            Player,
            SeasonPlayer,
            Team,
            TeamMember,
        )

        teams_by_name = {}
        for number, team_name in enumerate(all_teams, start=1):
            roster_rows = rosters.get(team_name, [])
            with reversion.create_revision():
                reversion.set_comment("Imported knockout team roster.")
                team = Team.objects.create(
                    season=season, number=number, name=team_name
                )

                ratings = []
                for row in roster_rows:
                    entry = plan[(team_name, row["board_number"])]
                    player = self._materialise_player(Player, entry)

                    if player.rating is not None:
                        ratings.append(player.rating)

                    SeasonPlayer.objects.get_or_create(
                        season=season,
                        player=player,
                        defaults={"is_active": True},
                    )
                    TeamMember.objects.create(
                        team=team,
                        player=player,
                        board_number=row["board_number"],
                        is_captain=row["role"] == "captain",
                        is_vice_captain=row["role"] == "vice_captain",
                    )

                if ratings:
                    team.seed_rating = round(sum(ratings) / len(ratings))
                    team.save()

            teams_by_name[team_name] = team
        return teams_by_name

    def _materialise_player(self, Player, entry):
        """Create (or fetch) the Player for a plan entry and backfill blanks."""
        player = entry["player"]
        if player is None:
            player = Player(lichess_username=entry["username"])
            if entry["fide_id"]:
                player.fide_id = entry["fide_id"]
            if entry["rating"] is not None:
                player.rating = entry["rating"]
            player.save()
            return player

        # Existing player: fill in blanks, never overwrite real data.
        changed = False
        if not player.fide_id and entry["fide_id"]:
            player.fide_id = entry["fide_id"]
            changed = True
        if player.rating is None and entry["rating"] is not None:
            player.rating = entry["rating"]
            changed = True
        if changed:
            player.save()
        return player

    def _create_bracket(self, season, bracket_size, match_generation):
        from heltour.tournament.models import KnockoutBracket

        bracket, _created = KnockoutBracket.objects.get_or_create(
            season=season,
            defaults={
                "bracket_size": bracket_size,
                "seeding_style": "traditional",
                "games_per_match": 1,
                "matches_per_stage": 2,
            },
        )
        # These are authoritative even on a re-run.
        bracket.match_generation = match_generation
        bracket.bracket_size = bracket_size
        bracket.matches_per_stage = 2
        bracket.save()
        return bracket

    def _create_seedings(self, bracket, all_teams, teams_by_name):
        from heltour.tournament.models import KnockoutSeeding

        for seed_number, team_name in enumerate(all_teams, start=1):
            KnockoutSeeding.objects.get_or_create(
                bracket=bracket,
                team=teams_by_name[team_name],
                defaults={"seed_number": seed_number, "is_manual_seed": True},
            )

    def _create_round1_pairings(self, round1, ordered_pairs, teams_by_name, bracket):
        import reversion

        from heltour.tournament.models import TeamPairing
        from heltour.tournament.pairinggen import (
            _create_board_pairings_for_knockout,
        )

        # In 'upfront' mode create both matches of the stage now; in 'lockstep'
        # mode only match 1. pairing_order is laid out match-by-match.
        matches_to_create = (
            bracket.matches_per_stage if bracket.match_generation == "upfront" else 1
        )
        total_pairs = len(ordered_pairs)
        created = 0

        for match_number in range(1, matches_to_create + 1):
            for pair_index, (_mn, a_name, b_name) in enumerate(ordered_pairs):
                team_a = teams_by_name[a_name]
                team_b = teams_by_name[b_name]
                # Match 1 keeps the bracket-file colors; match 2 swaps them.
                if match_number % 2 == 1:
                    white_team, black_team = team_a, team_b
                else:
                    white_team, black_team = team_b, team_a

                pairing_order = (match_number - 1) * total_pairs + pair_index + 1

                with reversion.create_revision():
                    reversion.set_comment("Imported knockout round-1 pairing.")
                    team_pairing = TeamPairing.objects.create(
                        white_team=white_team,
                        black_team=black_team,
                        round=round1,
                        pairing_order=pairing_order,
                    )

                _create_board_pairings_for_knockout(team_pairing, round1.season.boards)
                created += 1

        return created

    # ----- parsing helpers ----------------------------------------------------

    @staticmethod
    def _find_header(rows):
        """Return (header_row_index, lowercased_header_cells) or (None, [])."""
        for idx, row in enumerate(rows):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if ROSTER_HEADER_KEY in cells:
                return idx, cells
        return None, []

    @staticmethod
    def _cell_getter(col, row):
        """Return a get(column_name) accessor over a row, given a name->index map."""
        def get(name):
            i = col.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        return get

    @staticmethod
    def _row_is_blank(get):
        return all(
            v is None or str(v).strip() == ""
            for v in (
                get("team_name"),
                get("lichess_username"),
                get("fide_id"),
                get("board_number"),
            )
        )

    # ----- normalisation helpers ----------------------------------------------

    @staticmethod
    def _strip_time(name):
        """Drop a trailing kick-off time from a bracket team name."""
        return TIME_TAIL_RE.sub("", str(name).strip()).strip()

    @staticmethod
    def _normalize_username(raw):
        """Return (username, warning_or_None). Handles '@name' and profile URLs."""
        if raw is None:
            return "", None
        s = str(raw).strip()
        if not s:
            return "", None
        original = s
        m = re.search(r"lichess\.org/@/([^/?\s]+)", s, re.IGNORECASE)
        if m:
            s = m.group(1)
        s = s.strip().lstrip("@").strip()
        if " " in s:
            s = s.split()[0]
        warning = None
        if s != original.strip():
            warning = f"normalised lichess_username {original!r} -> {s!r}"
        return s, warning

    @staticmethod
    def _normalize_fide_id(raw):
        """Return (fide_id, warning_or_None). FIDE ids are digits only."""
        if raw is None:
            return "", None
        s = str(raw).strip()
        if s in ("", "0"):
            return "", None
        digits = re.sub(r"\s+", "", s)
        if digits.isdigit():
            return digits.lstrip("0") or digits, None
        return "", f"ignoring non-numeric fide_id {raw!r}"

    @staticmethod
    def _normalize_role(raw):
        """Return (role, warning_or_None). role is 'captain'/'vice_captain'/''."""
        if raw is None:
            return "", None
        s = str(raw).strip().lower()
        if not s:
            return "", None
        if s in CAPTAIN_ROLES:
            return "captain", None
        if s in VICE_CAPTAIN_ROLES:
            return "vice_captain", None
        return "", f"unrecognised role {raw!r}; treating as a regular player"

    @staticmethod
    def _same_username(a, b):
        return (a or "").strip().lower() == (b or "").strip().lower()

    @staticmethod
    def _fold(name):
        return re.sub(r"\s+", " ", str(name).strip()).casefold()

    @staticmethod
    def _parse_int(value):
        if value is None:
            return None
        s = str(value).strip().replace(",", "")
        if s == "":
            return None
        try:
            return int(float(s))
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _is_power_of_2(n):
        return n > 0 and (n & (n - 1)) == 0
